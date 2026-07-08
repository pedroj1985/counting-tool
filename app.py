import streamlit as st
import pandas as pd
import re
from io import BytesIO

st.set_page_config(page_title="counting-tool", layout="centered", initial_sidebar_state="collapsed")

st.markdown("""
<style>
    .block-container { padding-top: 1.5rem; padding-bottom: 1.5rem; }
    .stExpander { border: 1px solid #e2e8f0; border-radius: 12px; margin-bottom: 1rem; }
    div[data-testid="stFileUploader"] { width: 100%; }
    .kpi-card {
        background: #ffffff;
        border: 1px solid #e2e8f0;
        border-radius: 12px;
        padding: 1.2rem 0.8rem;
        text-align: center;
        box-shadow: 0 1px 3px rgba(0,0,0,0.04);
    }
    .kpi-card .kpi-value { font-size: 1.8rem; font-weight: 700; line-height: 1.2; }
    .kpi-card .kpi-label { font-size: 0.8rem; color: #64748b; margin-top: 0.25rem; }
    .kpi-card.good { border-top: 4px solid #10b981; }
    .kpi-card.total { border-top: 4px solid #2563eb; }
    .kpi-card.warn { border-top: 4px solid #f59e0b; }
    .kpi-card.bad { border-top: 4px solid #ef4444; }
    .tag-ok { background: #d1fae5; color: #065f46; padding: 2px 10px; border-radius: 999px; font-size: 0.75rem; font-weight: 600; white-space: nowrap; }
    .tag-miss { background: #fee2e2; color: #991b1b; padding: 2px 10px; border-radius: 999px; font-size: 0.75rem; font-weight: 600; white-space: nowrap; }
    .tag-extra { background: #fef3c7; color: #92400e; padding: 2px 10px; border-radius: 999px; font-size: 0.75rem; font-weight: 600; white-space: nowrap; }
    div[data-testid="stMetricValue"] { font-size: 1.8rem !important; }
    .stAlert { border-radius: 10px; }
    .stDataFrame { border: 1px solid #e2e8f0; border-radius: 10px; overflow: hidden; }
    h1, h2, h3 { color: #1e293b; }
    hr { margin: 0.5rem 0; opacity: 0.3; }
</style>
""", unsafe_allow_html=True)

st.title("📦 counting-tool")
st.caption("Comparación de Manifiesto vs Conteo de Bultos")

# ── Session state ──────────────────────────────────────────────
if "manifest_df" not in st.session_state:
    st.session_state.manifest_df = None
    st.session_state.manifest_meta = {}
if "conteo1_series" not in st.session_state:
    st.session_state.conteo1_series = None
    st.session_state.conteo2_series = None
    st.session_state.conteo1_norm = None
    st.session_state.conteo2_norm = None
    st.session_state.conteo1_file = None
    st.session_state.conteo2_file = None

# ── Helpers ────────────────────────────────────────────────────

def normalize_bl(raw):
    raw = str(raw).strip()
    if not raw:
        return None
    raw = raw.upper()
    raw = re.sub(r'\s+', ' ', raw)
    m = re.match(r'^(CPK)[=\s\-]?\s*([A-Z0-9]+)', raw)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    m = re.match(r'^(K)[=\s\-]?\s*([A-Z0-9]+)', raw)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    m = re.match(r'^([A-Z]+)[=\s\-]?\s*(\d+)', raw)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    return raw


def parse_number(val):
    if val is None:
        return None
    try:
        return float(str(val).replace(",", "").strip())
    except:
        return None


def load_manifest(file):
    ext = file.name.rsplit(".", 1)[-1].lower()
    raw = file.read()

    if ext == "xls":
        import xlrd
        wb = xlrd.open_workbook(file_contents=raw)
        ws = wb.sheets()[0]
        rows = []
        for r in range(ws.nrows):
            rows.append([ws.cell_value(r, c) for c in range(ws.ncols)])
    else:
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(raw), data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([v if v is not None else "" for v in row])

    meta = {}
    for r in rows[:11]:
        label = str(r[0]).strip() if len(r) > 0 else ""
        value = str(r[1]).strip() if len(r) > 1 else ""
        if "MBL" in label or "MAWB" in label:
            meta["mbl"] = value
        elif "Contenedor" in label or "Container" in label:
            meta["container"] = value
        elif "Cantidad de Env" in label:
            meta["envios"] = parse_number(value)
        elif "Cantidad de Bultos" in label:
            meta["bultos"] = parse_number(value)
        elif "Peso" in label:
            meta["peso"] = parse_number(value)
        elif "Fecha" in label:
            meta["fecha"] = value
        elif "Agencia" in label or "Origen" in label:
            meta["agencia"] = value

    data_start = 11
    header_row = rows[data_start] if data_start < len(rows) else []
    data_rows = rows[data_start + 1:] if data_start + 1 < len(rows) else []

    df = pd.DataFrame(data_rows)
    if len(header_row) > 0:
        header_clean = [str(h).strip() for h in header_row]
        df.columns = header_clean[:df.shape[1]] if df.shape[1] <= len(header_clean) else header_clean + [""] * (df.shape[1] - len(header_clean))

    if "Número de BL" in df.columns:
        df = df[df["Número de BL"].astype(str).str.upper().str.strip() != "TOTAL"]
        df = df[df["Número de BL"].astype(str).str.strip() != ""]
        df = df.reset_index(drop=True)

    return df, meta


def load_conteo(file):
    raw = file.read()
    ext = file.name.rsplit(".", 1)[-1].lower()
    if ext == "xls":
        import xlrd
        wb = xlrd.open_workbook(file_contents=raw)
        ws = wb.sheets()[0]
        vals = [str(ws.cell_value(r, 0)).strip() for r in range(ws.nrows)]
    else:
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(raw), data_only=True)
        ws = wb[wb.sheetnames[0]]
        vals = [str(row[0]).strip() if row[0] else "" for row in ws.iter_rows(values_only=True)]
    vals = [v for v in vals if v and v != "None" and v != ""]
    return pd.Series(vals, name="BL")


def get_manifest_bl_set(df):
    if "Número de BL" not in df.columns:
        return set()
    bls = df["Número de BL"].astype(str).str.strip()
    norm = bls.apply(normalize_bl)
    return set(norm.dropna().tolist())


# ────────────────────────────────────────────────────────────────
# SECTION 1 — MANIFIESTO
# ────────────────────────────────────────────────────────────────
with st.expander("**1. Manifiesto**", expanded=True):
    uploaded_manifiesto = st.file_uploader(
        "Subir archivo del Manifiesto",
        type=["xls", "xlsx"],
        key="manifiesto_uploader",
        label_visibility="collapsed",
        help="Selecciona el archivo .xls o .xlsx del manifiesto"
    )

    if uploaded_manifiesto is not None:
        try:
            df, meta = load_manifest(uploaded_manifiesto)
            st.session_state.manifest_df = df
            st.session_state.manifest_meta = meta
        except Exception as e:
            st.error(f"Error al leer el manifiesto: {e}")

    if st.session_state.manifest_df is not None:
        df = st.session_state.manifest_df
        meta = st.session_state.manifest_meta

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            val = meta.get("mbl", meta.get("container", "—"))
            st.metric("MBL / Contenedor", val if len(val) < 20 else val[:18] + "…")
        with c2:
            env = int(meta["envios"]) if meta.get("envios") else len(df)
            st.metric("Total Envíos", env)
        with c3:
            bul = int(meta["bultos"]) if meta.get("bultos") else "—"
            st.metric("Total Bultos", bul)
        with c4:
            peso = meta.get("peso", "—")
            if isinstance(peso, (int, float)):
                st.metric("Peso Total (kg)", f"{peso:,.2f}")
            else:
                st.metric("Peso Total (kg)", peso)

        search = st.text_input("🔍 Buscar por BL, destinatario o provincia", placeholder="Escribe para filtrar…", label_visibility="collapsed")

        display_cols = ["Número de BL", "Bultos", "Peso Kg", "Nombre y Apellidos del Destinatario", "Provincia", "Municipio", "Descripcion de las Mercancias"]
        available_cols = [c for c in display_cols if c in df.columns]
        df_display = df[available_cols].copy()

        for col in df_display.columns:
            df_display[col] = df_display[col].astype(str)

        if search:
            mask = df_display.apply(lambda x: x.str.contains(search, case=False, na=False)).any(axis=1)
            df_display = df_display[mask]

        st.dataframe(
            df_display,
            width='stretch',
            height=min(320, 40 * (len(df_display) + 1)),
            hide_index=True,
        )
        st.caption(f"Mostrando {len(df_display)} de {len(df)} registros")

# ────────────────────────────────────────────────────────────────
# SECTION 2 — CONTEOS
# ────────────────────────────────────────────────────────────────
with st.expander("**2. Conteos**", expanded=True):
    col1, col2 = st.columns(2)
    with col1:
        c1_file = st.file_uploader("Conteo 1", type=["xls", "xlsx"], key="conteo1", label_visibility="collapsed")
    with col2:
        c2_file = st.file_uploader("Conteo 2", type=["xls", "xlsx"], key="conteo2", label_visibility="collapsed")

    results = []
    for f, key, key_norm, key_file, label in [
        (c1_file, "conteo1_series", "conteo1_norm", "conteo1_file", "Conteo 1"),
        (c2_file, "conteo2_series", "conteo2_norm", "conteo2_file", "Conteo 2"),
    ]:
        if f is not None:
            if st.session_state.get(key_file) != f.name:
                try:
                    series = load_conteo(f)
                    st.session_state[key] = series
                    st.session_state[key_norm] = series.apply(normalize_bl).dropna()
                    st.session_state[key_file] = f.name
                except Exception as e:
                    st.error(f"Error en {label}: {e}")

        s = st.session_state.get(key)
        if s is not None and not s.empty:
            n = st.session_state[key_norm]
            results.append((label, len(s), n))

    for label, raw_count, normalized in results:
        st.markdown(f"**{label}**: {raw_count} entradas")

    if st.session_state.conteo1_norm is not None and st.session_state.conteo2_norm is not None:
        union = set(st.session_state.conteo1_norm.tolist()) | set(st.session_state.conteo2_norm.tolist())
        st.info(f"🧮 **Total combinado** (unicos): {len(union)} BLs")
    elif st.session_state.conteo1_norm is not None:
        st.info(f"🧮 **Total**: {len(st.session_state.conteo1_norm)} BLs")
    elif st.session_state.conteo2_norm is not None:
        st.info(f"🧮 **Total**: {len(st.session_state.conteo2_norm)} BLs")

# ────────────────────────────────────────────────────────────────
# SECTION 3 — DASHBOARD
# ────────────────────────────────────────────────────────────────
with st.expander("**3. Dashboard**", expanded=True):
    if st.session_state.manifest_df is None:
        st.warning("Carga primero el Manifiesto en la sección 1.")
    else:
        df = st.session_state.manifest_df
        manifest_bls = get_manifest_bl_set(df)
        total_bl = len(manifest_bls)

        conteo_bls = set()
        if st.session_state.conteo1_norm is not None:
            conteo_bls |= set(st.session_state.conteo1_norm.tolist())
        if st.session_state.conteo2_norm is not None:
            conteo_bls |= set(st.session_state.conteo2_norm.tolist())

        if not conteo_bls:
            st.warning("Carga al menos un Conteo en la sección 2.")
        else:
            contados = manifest_bls & conteo_bls
            faltantes = manifest_bls - conteo_bls
            extraños = conteo_bls - manifest_bls

            pct = round(len(contados) / total_bl * 100, 1) if total_bl else 0

            st.markdown("### Resumen")
            kcols = st.columns(3)
            with kcols[0]:
                st.markdown(f'<div class="kpi-card total"><div class="kpi-value">{total_bl}</div><div class="kpi-label">Total BL en Manifiesto</div></div>', unsafe_allow_html=True)
            with kcols[1]:
                cls = "good" if pct >= 80 else ("warn" if pct >= 50 else "bad")
                st.markdown(f'<div class="kpi-card {cls}"><div class="kpi-value">{len(contados)}</div><div class="kpi-label">BL Contados</div></div>', unsafe_allow_html=True)
            with kcols[2]:
                cls = "good" if pct >= 80 else ("warn" if pct >= 50 else "bad")
                st.markdown(f'<div class="kpi-card {cls}"><div class="kpi-value">{pct}%</div><div class="kpi-label">Cobertura</div></div>', unsafe_allow_html=True)

            st.markdown("### Detalle por BL")
            rows = []
            for bl in sorted(manifest_bls):
                status = "✅ Contado" if bl in contados else "❌ Faltante"
                rows.append({"BL": bl, "Estado": status})
            for bl in sorted(extraños):
                rows.append({"BL": bl, "Estado": "⚠️ Extraño"})

            detail_df = pd.DataFrame(rows)
            st.dataframe(
                detail_df,
                column_order=["BL", "Estado"],
                width='stretch',
                height=min(360, 35 * (len(detail_df) + 1)),
                hide_index=True,
            )

            st.markdown("### Detalle de Conteos")

            rc1 = len(st.session_state.conteo1_norm) if st.session_state.conteo1_norm is not None else 0
            rc2 = len(st.session_state.conteo2_norm) if st.session_state.conteo2_norm is not None else 0
            rcat = len(conteo_bls)

            cc = st.columns(4)
            labels = ["Conteo 1", "Conteo 2", "Combinados", "BL Matcheados"]
            vals = [rc1, rc2, rcat, len(contados)]
            for i, (l, v) in enumerate(zip(labels, vals)):
                with cc[i]:
                    st.metric(l, v)

            if extraños:
                extra_str = ", ".join(sorted(extraños))
                st.markdown(f'<div style="margin-top:0.5rem;"><span class="tag-extra">⚠️ {len(extraños)} BL extraños no encontrados en manifiesto</span></div>', unsafe_allow_html=True)
                with st.expander("Ver BLs extraños"):
                    st.markdown(extra_str)

            if faltantes:
                falt_str = ", ".join(sorted(faltantes))
                st.markdown(f'<div style="margin-top:0.5rem;"><span class="tag-miss">❌ {len(faltantes)} BL faltantes por contar</span></div>', unsafe_allow_html=True)
                with st.expander("Ver BLs faltantes"):
                    st.markdown(falt_str)

            if st.session_state.conteo1_norm is not None and st.session_state.conteo2_norm is not None:
                only1 = set(st.session_state.conteo1_norm.tolist()) - set(st.session_state.conteo2_norm.tolist())
                only2 = set(st.session_state.conteo2_norm.tolist()) - set(st.session_state.conteo1_norm.tolist())
                both = set(st.session_state.conteo1_norm.tolist()) & set(st.session_state.conteo2_norm.tolist())
                bc = st.columns(3)
                with bc[0]:
                    st.metric("Solo en Conteo 1", len(only1))
                with bc[1]:
                    st.metric("Solo en Conteo 2", len(only2))
                with bc[2]:
                    st.metric("En Ambos", len(both))


st.markdown("---")
st.caption("counting-tool · Procesamiento de Manifiesto vs Conteo de Bultos")
